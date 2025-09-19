"""
Google Drive의 Excel 파일을 직접 조작하는 서비스
"""
import io
import tempfile
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
import openpyxl
from openpyxl.utils import get_column_letter

class ExcelDriveService:
    def __init__(self, credentials_path, file_id):
        self.credentials_path = credentials_path
        self.file_id = file_id
        self.drive_service = self._build_drive_service()
    
    def _build_drive_service(self):
        creds = service_account.Credentials.from_service_account_file(
            self.credentials_path,
            scopes=['https://www.googleapis.com/auth/drive']
        )
        return build('drive', 'v3', credentials=creds)
    
    def update_excel_data(self, sheet_name, row_data):
        """Excel 파일 다운로드 -> 수정 -> 업로드"""
        
        # 1. 파일 다운로드
        request = self.drive_service.files().export_media(
            fileId=self.file_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        file_content = io.BytesIO()
        downloader = MediaIoBaseDownload(file_content, request)
        
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        
        file_content.seek(0)
        
        # 2. Excel 파일 수정
        workbook = openpyxl.load_workbook(file_content)
        
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        
        # 빈 행 찾기
        next_row = worksheet.max_row + 1
        
        # 데이터 입력 (I, L, N, O 열)
        columns = {'I': 9, 'L': 12, 'N': 14, 'O': 15}
        
        for col_letter, col_data in row_data.items():
            if col_letter in columns:
                col_num = columns[col_letter]
                worksheet.cell(row=next_row, column=col_num, value=col_data)
        
        # 3. 수정된 파일 업로드
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        media = MediaIoBaseUpload(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        self.drive_service.files().update(
            fileId=self.file_id,
            media_body=media
        ).execute()
        
        print(f"✅ Excel 파일 업데이트 완료 (행 {next_row})")
        return next_row
